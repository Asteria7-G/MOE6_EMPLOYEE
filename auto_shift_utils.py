#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import load_workbook
import pandas as pd
import math
from ortools.sat.python import cp_model
import numpy as np
from openpyxl import load_workbook
from datetime import datetime, timedelta
import pymssql
from sqlalchemy import create_engine


# In[36]:


def df_to_sql(df):
    df.reset_index(drop=True, inplace=True)
    print('***********', df.loc[0, 'ytdAveNewOT'])
    connect = pymssql.connect('szhaeapp113', 'OTUser', 'OTUser@123', 'OTReginster')
    cursor = connect.cursor()
    for i in range(len(df)):
        status = df.loc[i, 'Status']
        year = df.loc[i, 'Year']
        month = df.loc[i, 'Month']
        day = df.loc[i, 'Day']
        shift = df.loc[i, 'Shift']
        shift_code = df.loc[i, 'ShiftCode']
        line = df.loc[i, 'Line']
        line_id = df.loc[i, 'LineId']
        OJTCode = df.loc[i, 'OJTCode']
        PersonNo = df.loc[i, 'PersonNo']
        Name = df.loc[i, 'Name']
        SkillName = df.loc[i, 'SkillName']
        Sshift = df.loc[i, 'Sshift']
        SAPShiftCode = df.loc[i, 'SAPShiftCode']
        ytdAveNewOT = float(df.loc[i, 'ytdAveNewOT'])
        MonthOT = float(df.loc[i, 'MonthOT'])
        Gender = df.loc[i, 'Gender']

        cursor.execute("insert into OT_Result values ('{}', '{}', '{}','{}', '{}', '{}','{}', '{}', '{}','{}', '{}', '{}','{}', '{}', '{}','{}', '{}')".format(
            status,year,month,day,shift,shift_code,line, line_id,OJTCode,PersonNo,Name,SkillName,Sshift,SAPShiftCode,ytdAveNewOT,MonthOT,Gender))
        connect.commit()
    
    cursor.close()
    connect.close()
    print(f"DataFrame has been imported to SQL Server table")


# In[3]:


#线体信息
# line_support_group_dict = {'LG1':[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,31],'LG2':[22,23,24,25,26,27,28,29,30,32,33]}
# line_group_to_one_person_dict = {'OJT3124':{'0.14':[1,2,3,4,5,6],'0.1':[8,9,10,11,12,13,14,15,16,17,18]}}


# In[4]:


# line id dict （固定） 线体名对应的id
# r'.\Data\autoshift.xlsx'
def read_excel_line_id(file_path):
    line_id_df = pd.read_excel(file_path,sheet_name='线体id')
    line_id_dict = {}
    for i in range(len(line_id_df)):
        line_id_dict[line_id_df.loc[i, 'Line'].replace(' ', '').lower()] = line_id_df.loc[i, 'LineId']
    return line_id_dict


# In[5]:


# #线体互换比重的matrix（固定）
# def read_excel_matrix(file_path):
#     wb = load_workbook(file_path)
#     sheet = wb.active

#     matrix = {}

#     for row in range(4, sheet.max_row + 1):  # 从第二行开始读取数据
#         key_row = sheet.cell(row=row, column=2).value
#         for col in range(3, sheet.max_column -1):  # 从第二列开始读取数据
#             key_col = sheet.cell(row=3, column=col).value
#             value = sheet.cell(row=row, column=col).value
#             matrix[(key_row, key_col)] = value

#     return matrix


# In[6]:


# # 读取 Excel 文件中的矩阵数据
# excel_matrix = read_excel_matrix(r'.\Data\技能互通矩阵图 .xlsx')

# # 打印矩阵数据
# line_support_priority_matrix = {}
# for key, value in excel_matrix.items():
#     a, b = key[0], key[1]
#     a = a.replace(' ', '').lower()
#     b = b.replace(' ', '').lower()
#     if a in line_id_dict and b in line_id_dict:
#         line_support_priority_matrix[(line_id_dict[a], line_id_dict[b])] = value
#     elif a == 'sal4':
#         if b == 'sal4':
#             line_support_priority_matrix[(29, 29)] = value
#             line_support_priority_matrix[(30, 29)] = value
#             line_support_priority_matrix[(29, 30)] = value
#             line_support_priority_matrix[(30, 30)] = value
#         elif b == 'sal6':
#             line_support_priority_matrix[(29, 32)] = value
#             line_support_priority_matrix[(30, 32)] = value
#             line_support_priority_matrix[(29, 33)] = value
#             line_support_priority_matrix[(30, 33)] = value
#         else:
#             line_support_priority_matrix[(29, line_id_dict[b])] = value
#             line_support_priority_matrix[(30, line_id_dict[b])] = value
           
#     elif b == 'sal4':
#         if a not in ['sal4', 'sal6']:
#             line_support_priority_matrix[(line_id_dict[a], 29)] = value
#             line_support_priority_matrix[(line_id_dict[a], 30)] = value
#     elif a == 'sal6':
#         if b == 'sal4':
#             line_support_priority_matrix[(32, 29)] = value
#             line_support_priority_matrix[(33, 29)] = value
#             line_support_priority_matrix[(32, 30)] = value
#             line_support_priority_matrix[(33, 30)] = value
#         elif b == 'sal6':
#             line_support_priority_matrix[(32, 32)] = value
#             line_support_priority_matrix[(33, 32)] = value
#             line_support_priority_matrix[(32, 33)] = value
#             line_support_priority_matrix[(33, 33)] = value
#         else:
#             line_support_priority_matrix[(32, line_id_dict[b])] = value
#             line_support_priority_matrix[(33, line_id_dict[b])] = value
#     elif b == 'sal6':
#         if a not in ['sal4', 'sal6']:
#             line_support_priority_matrix[(line_id_dict[a], 32)] = value
#             line_support_priority_matrix[(line_id_dict[a], 33)] = value
#     else:
#         print('Wrong')
    
# line_support_priority_matrix


# In[7]:


def get_keys_from_value(my_dict, search_value):
    for key, value in\
            my_dict.items():
        if value == search_value:
            return key


# In[8]:


#选出这个group的人 不标准问题吧 SAL4和SAL6 会存在两条线 IB2FA1 是IB2FA01的非标准写法
def line2lineid(line,line_id_dict):
    if line not in ['SAL4', 'SAL6', 'IB2FA1', 'IB2FA2']:
        return line_id_dict[line.replace(' ', '').lower()]
    elif line == 'SAL4':
        return '29/30'
    elif line == 'SAL6':
        return '32/33'
    elif line == 'IB2FA1':
        return 20
    elif line == 'IB2FA2':
        return 21


# In[9]:


#人员信息 HC剔除line leader、选性别和本线体
# r'.\Data\HC.xlsx'
def read_excel_person_info(hc_file_path):
    person_info_1_df = pd.read_excel(hc_file_path,sheet_name='MFO6.1')
    person_info_2_df = pd.read_excel(hc_file_path,sheet_name='MFO6.2')
    person_info_df = pd.concat([person_info_1_df, person_info_2_df], ignore_index=True)
    ll_info_df = person_info_df[person_info_df['Position '].notna()]
    op_info_df = person_info_df[person_info_df['Position '].isna()]
    op_info_df.rename(columns={'Pers.No.':'PersonNo'},inplace=True)
    op_info_df = op_info_df[['PersonNo','Name','性别','Line']]
    op_info_df['PersonNo'] = op_info_df['PersonNo'].astype(int).astype(str)
    return ll_info_df, op_info_df


# In[10]:


# r'.\Data\ishopfloor_front_OT.xlsx'
def read_excel_ishopfloor_person_ot(ot_file_path):
    ishopfloor_ot_df = pd.read_excel(ot_file_path,sheet_name='OT')
    ot_df = ishopfloor_ot_df[['Pers.No.','当月OT','年度AveNew']]
    ot_df.rename(columns={'Pers.No.':'PersonNo'},inplace=True)
    ot_df['PersonNo'] = ot_df['PersonNo'].astype(int).astype(str)
    return ot_df


# In[11]:


#技能
# r'.\Data\1127_TQS.xlsx'
def read_excel_tqs_skill(skill_file_path):
    tqs_df = pd.read_excel(skill_file_path)
    tqs_df = tqs_df[['PersonNo','OJTCode','SkillName']]
    tqs_df['PersonNo'] = tqs_df['PersonNo'].astype(int).astype(str)
    return tqs_df


# In[12]:


#ishopfloor shift
# r'.\Data\ishopfloor_end_shift.xlsx'
def read_excel_ishopfloor_person_shift(shift_file_path):
    ishopfloor_shift_df = pd.read_excel(shift_file_path,sheet_name='Shift')
    ishopfloor_shift_df.dropna(subset=['PersonNo'],inplace=True)
    ishopfloor_shift_df['PersonNo'] = ishopfloor_shift_df['PersonNo'].astype(int).astype(str)
    return ishopfloor_shift_df


# In[13]:


def select_off_person_without_7_working_possibility(year, month, day, person_list, shift_df):
    """
    筛选出前后各有一次休息班且间隔不超过7天的人员
    并打印调试信息
    """
    off_code_list = {
        'roff','100','f200','p1800','l2100','q500','q600','q700','q800',
        'q900','q1000','q1200','f1100','s2700','g300',',c2600','p1701',
        'q901','q9400','p1900'
    }

    # 转成字典提高查询效率
    shift_dict = {}
    for idx, row in shift_df.iterrows():
        # {[personno,date]: sap_shift_code}
        key = (row['PersonNo'], datetime(row['Syear'], row['Smonth'], row['Sdate']).date())
        value = (str(row['Sshift']).strip().lower(), str(row['sap_shift_code']).strip().lower())
        shift_dict[key] = value

    def find_off_ay(person_no, start_date, days=6, direction='past'):
        for i in range(1, days + 1):
            date = start_date - timedelta(days=i) if direction == 'past' else start_date + timedelta(days=i)
            key = (person_no, date.date())
            if key in shift_dict:
                shift, code = shift_dict[key]
                combined_code = (shift + code).replace(' ', '').lower()
                if combined_cdode in off_code_list:
                    return date
        return None

    today = datetime(year, month, day)
    result_list = []

    print(f"===== 调试信息: {today.date()} =====")
    for person_no in person_list:
        past_off = find_off_day(person_no, today, direction='past')
        future_off = find_off_day(person_no, today, direction='future')

        if past_off:
            print(f"{person_no} 前休息班: {past_off.date()}")
        else:
            print(f"{person_no} 前休息班: 无")

        if future_off:
            print(f"{person_no} 后休息班: {future_off.date()}")
        else:
            print(f"{person_no} 后休息班: 无")

        if past_off and future_off:
            delta_days = (future_off - past_off).days
            print(f"{person_no} 前后休息班间隔: {delta_days} 天")
            # 判断间隔是否 ≤ 7 天
            if delta_days <= 7:
                result_list.append(person_no)
                print(f"--> {person_no} 符合条件 ✅")
            else:
                print(f"--> {person_no} 不符合间隔条件 ❌")
        print("-" * 40)

    print("最终符合条件人员:", result_list)
    return result_list


# In[14]:


def person_for_use(op_info_df,ot_df,tqs_df,ishopfloor_shift_df,year,month,day,shift,shift_code,line_id_dict,working_line_id_list,off_person_used_list):
    print('=============================off person used list =============================')
    print(off_person_used_list)
    # person表+ot col 将 人员信息表 (op_info_df) 和 OT 工时表 (ot_df) 合并
    op_info_df = pd.merge(op_info_df, ot_df, on='PersonNo', how='left')
    # 筛年度ot小于35 只保留 年度平均 OT 工时小于等于 35 的员工。
    op_info_df = op_info_df[op_info_df['年度AveNew']<=35]
    # 对应line name +line id col 把 Line 转换为 LineId，可能会出现一个人属于多个 line 的情况：
    # 如果 LineId 是字符串（例如 "1/2"），就把这个人复制一行，分别赋值成 1 和 2。
    op_info_df['LineId'] = op_info_df['Line'].apply(lambda x:line2lineid(x,line_id_dict))
    op_info_df.reset_index(drop=True, inplace=True)
    length = len(op_info_df)
    for i in range(length):
        if type(op_info_df.loc[i, 'LineId']) == str:
            line_id = op_info_df.loc[i, 'LineId'].split('/')
            op_info_df.loc[i, 'LineId'] = int(line_id[0])
            len_new = len(op_info_df)
            op_info_df.loc[len_new] = op_info_df.loc[i]
            op_info_df.loc[len_new, 'LineId'] = int(line_id[1])

    #filter这个groupline相关的员工  判断 working_line_id_list[0] 属于哪一组 (LG1/LG2)， 只保留这个组别里的员工。
    line_support_group_dict = {'LG1':[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,31],'LG2':[22,23,24,25,26,27,28,29,30,32,33]}
    for line_support_group_dict_value in line_support_group_dict.values():
        if working_line_id_list[0] in line_support_group_dict_value:
            filter_op_info_df = op_info_df[op_info_df['LineId'].isin(line_support_group_dict_value)]
    #         print(filter_op_info_df)
    #员工表加技能列
    filter_op_info_df['PersonNo'] = filter_op_info_df['PersonNo'].astype(int).astype(str)
    filter_op_info_df = pd.merge(filter_op_info_df, tqs_df, on='PersonNo', how='left')
    #         print(filter_op_info_df)
    #筛当天的排班
    filter_ishopfloor_shift_df = ishopfloor_shift_df[(ishopfloor_shift_df['Syear'] == year) & (ishopfloor_shift_df['Smonth'] == month) &(ishopfloor_shift_df['Sdate'] == day)]
#     filter_ishopfloor_shift_df['PersonNo'] = filter_ishopfloor_shift_df['PersonNo'].astype(int).astype(str)
    #         print(filter_ishopfloor_shift_df)
    filter_ishopfloor_shift_df = filter_ishopfloor_shift_df[filter_ishopfloor_shift_df['PersonNo'].isin(filter_op_info_df['PersonNo'].values)]
    if shift_code in ['S121','S080','S101','S084','S104']:
        shift_code = 'S121'
    elif shift_code in ['S122','S102','S083','S103','S082']:
        shift_code = 'S122'
    filter_ishopfloor_shift_df_a = filter_ishopfloor_shift_df[(filter_ishopfloor_shift_df['Sshift'] == shift) & (filter_ishopfloor_shift_df['sap_shift_code'] == shift_code)]
    filter_ishopfloor_shift_df_b = filter_ishopfloor_shift_df[(filter_ishopfloor_shift_df['Sshift'] == 'R') & (filter_ishopfloor_shift_df['sap_shift_code'] == 'OFF')]
    filter_ishopfloor_shift_df = pd.concat([filter_ishopfloor_shift_df_a, filter_ishopfloor_shift_df_b], ignore_index=True)
    filter_ishopfloor_shift_df = filter_ishopfloor_shift_df[['PersonNo', 'Sshift','sap_shift_code']]
    person_ishopfloor_shift_df = filter_ishopfloor_shift_df.merge(filter_op_info_df, on='PersonNo', how='left')
    #筛选出来的人再分成正常班和off两个df，要对off的人做处理，删除off转成working会导致连续七天上班的人
    person_ishopfloor_shift_working_df = person_ishopfloor_shift_df[(person_ishopfloor_shift_df['Sshift'] == shift) & (person_ishopfloor_shift_df['sap_shift_code'] == shift_code)]
    person_ishopfloor_shift_off_df = person_ishopfloor_shift_df[(person_ishopfloor_shift_df['Sshift'] == 'R') & (person_ishopfloor_shift_df['sap_shift_code'] == 'OFF')]
    person_ishopfloor_shift_off_list = list(set(person_ishopfloor_shift_off_df['PersonNo']))
    person_off_keep_waiting_list = select_off_person_without_7_working_possibility(year,month,day,person_ishopfloor_shift_off_list,ishopfloor_shift_df)
    filter_person_ishopfloor_shift_off_df = person_ishopfloor_shift_off_df[person_ishopfloor_shift_off_df['PersonNo'].isin(person_off_keep_waiting_list)]
    if off_person_used_list:
        filter_person_ishopfloor_shift_off_df = filter_person_ishopfloor_shift_off_df[~filter_person_ishopfloor_shift_off_df['PersonNo'].isin(off_person_used_list)]
    person_for_use_df = pd.concat([person_ishopfloor_shift_working_df, filter_person_ishopfloor_shift_off_df], ignore_index=True)
    
#     print(person_for_use_df[(person_for_use_df['PersonNo'] == '88958365') & (person_for_use_df['OJTCode'] == 'OJT3116')])
    return person_for_use_df


# In[15]:


#开班数据从数据库获取 {"['Year','Month','Day','DorN','ShiftCode']":[[LG1的开班线体list],[LG2的开班线体list]],....}
def read_excel_working_line():
    line_support_group_dict = {'LG1':[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,31],'LG2':[22,23,24,25,26,27,28,29,30,32,33]}
    connect = pymssql.connect(server='szhaeapp113',
                           user='OTUser', password='OTUser@123', database='OTReginster')
    cursor = connect.cursor() 
    cursor.execute("select * from OT_Plan")
    cols = [i[0] for i in cursor.description]
    values = cursor.fetchall()
    working_line_df = pd.DataFrame(data=values, columns=cols) 
    working_dict = {}
    working_line_df_group = working_line_df.groupby(['Year','Month','Day','DorN','ShiftCode'])
    for key, value in working_line_df_group:
        lg_list = []
        lg1_list = []
        lg2_list = []
        if key[4] != 'OFF':
            for line_id in list(value['LineId']):
                if int(line_id) in line_support_group_dict['LG1']:
                    lg1_list.append(int(line_id))
                elif int(line_id) in line_support_group_dict['LG2']:
                    lg2_list.append(int(line_id))
            if lg1_list:
                lg_list.append(lg1_list)
            if lg2_list:
                lg_list.append(lg2_list)
            working_dict[key] = lg_list
    return working_dict


# In[16]:


read_excel_working_line()


# In[17]:


# # 获取开班数据转换成dictionary key为时间+班次 value为list包含linegroup
# # r'.\Data\autoshift.xlsx'
# def read_excel_working_line(file_path):
#     line_support_group_dict = {'LG1':[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,31],'LG2':[22,23,24,25,26,27,28,29,30,32,33]}
#     working_line_df = pd.read_excel(file_path,sheet_name='开班')
#     working_dict = {}
#     working_line_df_group = working_line_df.groupby(['Year','Month','Date','DorN','ShiftCode'])
#     for key, value in working_line_df_group:
#         lg_list = []
#         lg1_list = []
#         lg2_list = []
#         if key[4] != 'OFF':
#             for line_id in list(value['LineId']):
#                 if line_id in line_support_group_dict['LG1']:
#                     lg1_list.append(line_id)
#                 elif line_id in line_support_group_dict['LG2']:
#                     lg2_list.append(line_id)
#             if lg1_list:
#                 lg_list.append(lg1_list)
#             if lg2_list:
#                 lg_list.append(lg2_list)
#             working_dict[key] = lg_list
#     return working_dict
# working_line_dict = read_excel_working_line(r'.\Data\autoshift.xlsx')
# working_line_dict


# In[18]:


# 线体技能所需要的人数（固定）
# r'.\Data\autoshift.xlsx'
def read_excel_line_need_skill_hc(file_path):
    line_skill_prefer_gender_df= pd.read_excel(file_path,sheet_name='线体技能需求人数')
    line_skill_prefer_gender_df_group = line_skill_prefer_gender_df.groupby(['LineId','OJTCode','Prefer'])
    line_skill_prefer_regular_dict = {}
    for key,value in line_skill_prefer_gender_df_group:
        if list(value['R'])[0] != 0:
            line_skill_prefer_regular_dict[key] = list(value['R'])[0]
    return line_skill_prefer_regular_dict


# In[19]:


# 获取本班次需要的技能和人头，有些线体是多条线组合公用一个人头
def working_line_prefer_regular_hc(line_skill_prefer_regular_dict,working_line_id_list):
    line_group_to_one_person_dict = {'OJT3124':{'0.14':[1,2,3,4,5,6],'0.1':[8,9,10,11,12,13,14,15,16,17,18]}}
    filter_line_group_to_one_person_dict = {}
    line_skill_perfer_to_hc_dict = {}
    for key in line_skill_prefer_regular_dict.keys():
        if key[0] in working_line_id_list:
            # 属于组合人头逻辑
            if line_skill_prefer_regular_dict[key] < 1:
                #组合技能是什么
                decimal_skill_name = key[1]
                if decimal_skill_name not in filter_line_group_to_one_person_dict:
                    filter_line_group_to_one_person_dict[decimal_skill_name] = {}
                #用组合技能来获取小数值，和小数值对应的线体
                for decimal_group in line_group_to_one_person_dict[decimal_skill_name].keys():
                    # 判断目前的线体在哪个小数值对应的线体组合里
                    if key[0] in line_group_to_one_person_dict[decimal_skill_name][decimal_group]:
                        # 看这个小数值是否已经存储过了
                        if decimal_group not in filter_line_group_to_one_person_dict[decimal_skill_name]:
                            filter_line_group_to_one_person_dict[decimal_skill_name][decimal_group] = {'line': [key[0]], 'gender':key[2]}
                        else:
                            filter_line_group_to_one_person_dict[decimal_skill_name][decimal_group]['line'].append(key[0])
            # 属于整数固定人头逻辑
            else:
                line_skill_perfer_to_hc_dict[key] = line_skill_prefer_regular_dict[key]
    for skill_key in filter_line_group_to_one_person_dict.keys():
        for decimal_key in filter_line_group_to_one_person_dict[skill_key]:
            gender = filter_line_group_to_one_person_dict[skill_key][decimal_key]['gender']
            line_skill_perfer_to_hc_dict[(decimal_key.replace('.', ''), skill_key, gender)] = math.ceil(len(filter_line_group_to_one_person_dict[skill_key][decimal_key]['line']) * float(decimal_key.split('/')[0]))
    return filter_line_group_to_one_person_dict,line_skill_perfer_to_hc_dict    


# In[20]:


def auto_shift(year,month,day,shift,shift_code,line_skill_perfer_to_hc_dict,filter_line_group_to_one_person_dict,person_for_use_df,working_line_name_list_string):
    model = cp_model.CpModel()

    weight_dict = {}
    variables_dict = {}
    for key in line_skill_perfer_to_hc_dict.keys():
        sub_variable_dict = {}
        all_person_no = []
        for person_no, group in person_for_use_df.groupby('PersonNo'):
            weight = 0
            if key[1] in list(group['OJTCode'].values) and group['性别'].values[0] in key[2]:
                all_person_no.append(person_no)
                # 动态创建变量并存储在字典中
                variables_dict[(key[0], key[1], person_no)] = model.NewBoolVar(f'{key[0]}_{key[1]}_{person_no}')
                sub_variable_dict[(key[0], key[1], person_no)] = variables_dict[(key[0], key[1], person_no)]

                if group['sap_shift_code'].values[0] == 'OFF':
                    weight += 1000000000
                if str(key[0])[0] != '0':
                    if len(list(group['LineId'])) == 1:
                        if group['LineId'].values[0] != key[0]:
                            weight += 10
                    #主要针对SAL4 和 SAL6 会有两个分支
                    elif len(list(group['LineId'])) == 2:
                        if group['LineId'].values[0] != key[0] and group['LineId'].values[1] != key[0]:
                            weight += 10
                        
                else:
                    line_list = filter_line_group_to_one_person_dict[key[1]][key[0][0]+'.'+key[0][1:]]['line']
                    if group['LineId'].values[0] not in line_list:
                        weight += 10
                if '/' in key[2] and group['性别'].values[0] == key[2][-1]:
                    weight += 1

                weight += int(group['年度AveNew'].values[0])*10000000
                weight += int(group['当月OT'].values[0])*100
                weight_dict[(key[0], key[1], person_no)] = weight

        model.Add(sum(sub_variable_dict[(key[0], key[1], p)] for p in all_person_no) == int(line_skill_perfer_to_hc_dict[key]))

    for person_no, group in person_for_use_df.groupby('PersonNo'):
        model.Add(sum(variables_dict[key] for key in variables_dict.keys() if key[2] == person_no) <= 1)


    model.Minimize(sum(variables_dict[key] * weight_dict[key] for key in variables_dict.keys()))

    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    if status == cp_model.OPTIMAL:
        print('*** Optimal solution found ***')
#         print([f'{key} = {solver.Value(variables_dict[key])}'for key in variables_dict.keys()])
        result_dict = {}
        for key in variables_dict.keys():
            result = f'{solver.Value(variables_dict[key])}'
            if result == '1' and str(key[0])[0] != '0':
                if 'status' not in result_dict:
                    result_dict['status'] = []
                    result_dict['status'].append('Optimal')
                else:
                    result_dict['status'].append('Optimal')
                if 'year' not in result_dict:
                    result_dict['year'] = []
                    result_dict['year'].append(year)
                else:
                    result_dict['year'].append(year)
                if 'month' not in result_dict:
                    result_dict['month'] = []
                    result_dict['month'].append(month)
                else:
                    result_dict['month'].append(month)
                if 'date' not in result_dict:
                    result_dict['date'] = []
                    result_dict['date'].append(day)
                else:
                    result_dict['date'].append(day)
                if 'shift' not in result_dict:
                    result_dict['shift'] = []
                    result_dict['shift'].append(shift)
                else:
                    result_dict['shift'].append(shift)
                if 'shift_code' not in result_dict:
                    result_dict['shift_code'] = []
                    result_dict['shift_code'].append(shift_code)
                else:
                    result_dict['shift_code'].append(shift_code)
                if 'line_id' not in result_dict:
                    result_dict['line_id'] = []
                    result_dict['line_id'].append(key[0])
                else:
                    result_dict['line_id'].append(key[0])
                if 'OJT_code'not in result_dict:
                    result_dict['OJT_code'] = []
                    result_dict['OJT_code'].append(key[1])
                else:
                    result_dict['OJT_code'].append(key[1])
                if 'person_no' not in result_dict:
                    result_dict['person_no'] = []
                    result_dict['person_no'].append(key[2])
                else:
                    result_dict['person_no'].append(key[2])   
            elif result == '1' and str(key[0])[0] == '0':
                line_list = filter_line_group_to_one_person_dict[key[1]][key[0][0]+'.'+key[0][1:]]['line']
                for l in line_list:
                    if 'status' not in result_dict:
                        result_dict['status'] = []
                        result_dict['status'].append('Optimal')
                    else:
                        result_dict['status'].append('Optimal')
                    if 'year' not in result_dict:
                        result_dict['year'] = []
                        result_dict['year'].append(year)
                    else:
                        result_dict['year'].append(year)
                    if 'month' not in result_dict:
                        result_dict['month'] = []
                        result_dict['month'].append(month)
                    else:
                        result_dict['month'].append(month)
                    if 'date' not in result_dict:
                        result_dict['date'] = []
                        result_dict['date'].append(day)
                    else:
                        result_dict['date'].append(day)
                    if 'shift' not in result_dict:
                        result_dict['shift'] = []
                        result_dict['shift'].append(shift)
                    else:
                        result_dict['shift'].append(shift)
                    if 'shift_code' not in result_dict:
                        result_dict['shift_code'] = []
                        result_dict['shift_code'].append(shift_code)
                    else:
                        result_dict['shift_code'].append(shift_code)
                    if 'line_id' not in result_dict:
                        result_dict['line_id'] = []
                        result_dict['line_id'].append(l)
                    else:
                        result_dict['line_id'].append(l)
                    if 'OJT_code'not in result_dict:
                        result_dict['OJT_code'] = []
                        result_dict['OJT_code'].append(key[1])
                    else:
                        result_dict['OJT_code'].append(key[1])
                    if 'person_no' not in result_dict:
                        result_dict['person_no'] = []
                        result_dict['person_no'].append(key[2])
                    else:
                        result_dict['person_no'].append(key[2])
                
        return 'optimal', 'Optimal solution found in %s - %s of group (%s) on %s/%s/%s'  % (shift,shift_code,working_line_name_list_string,year,month,day), result_dict


    else:
        print('*** No solution found ***')
        return 'nosolution','No solution found in %s - %s of group (%s) on %s/%s/%s' % (shift,shift_code,working_line_name_list_string,year,month,day),[]


# In[21]:


def auto_shift_no_solution(year,month,day,shift,shift_code,line_skill_perfer_to_hc_dict,filter_line_group_to_one_person_dict,person_for_use_df,working_line_name_list_string):
    model = cp_model.CpModel()

    weight_dict = {}
    variables_dict = {}
    working_position_count = 0
    for key in line_skill_perfer_to_hc_dict.keys():
        sub_variable_dict = {}
        all_person_no = []
        for person_no, group in person_for_use_df.groupby('PersonNo'):
            weight = 0
            if key[1] in list(group['OJTCode'].values) and group['性别'].values[0] in key[2]:
                all_person_no.append(person_no)
                # 动态创建变量并存储在字典中
                variables_dict[(key[0], key[1], person_no)] = model.NewBoolVar(f'{key[0]}_{key[1]}_{person_no}')
                sub_variable_dict[(key[0], key[1], person_no)] = variables_dict[(key[0], key[1], person_no)]

                if group['sap_shift_code'].values[0] == 'OFF':
                    weight += 1000000000
                if str(key[0])[0] != '0':
                    if len(list(group['LineId'])) == 1:
                        if group['LineId'].values[0] != key[0]:
                            weight += 10
                    #主要针对SAL4 和 SAL6 会有两个分支
                    elif len(list(group['LineId'])) == 2:
                        if group['LineId'].values[0] != key[0] and group['LineId'].values[1] != key[0]:
                            weight += 10
                        
                else:
                    line_list = filter_line_group_to_one_person_dict[key[1]][key[0][0]+'.'+key[0][1:]]['line']
                    if group['LineId'].values[0] not in line_list:
                        weight += 10
                if '/' in key[2] and group['性别'].values[0] == key[2][-1]:
                    weight += 1

                weight += int(group['年度AveNew'].values[0])*10000000
                weight += int(group['当月OT'].values[0])*100
                weight_dict[(key[0], key[1], person_no)] = weight

        model.Add(sum(sub_variable_dict[(key[0], key[1], p)] for p in all_person_no) <= int(line_skill_perfer_to_hc_dict[key]))
        working_position_count += sum(sub_variable_dict[(key[0], key[1], p)] for p in all_person_no)
        
    for person_no, group in person_for_use_df.groupby('PersonNo'):
        model.Add(sum(variables_dict[key] for key in variables_dict.keys() if key[2] == person_no) <= 1)
        
    working_position_count_weight = working_position_count*10000000000


    model.Maximize(working_position_count_weight - sum(variables_dict[key] * weight_dict[key] for key in variables_dict.keys()))

    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    if status == cp_model.OPTIMAL:
        print('*** Optimal solution found in round two ***')
#         print([f'{key} = {solver.Value(variables_dict[key])}'for key in variables_dict.keys()])
        result_dict = {}
        compare_no_solution_cause_dict_end = {}
        for key in variables_dict.keys():
            result = f'{solver.Value(variables_dict[key])}'  
            if result == '1':
                if (key[0],key[1]) not in compare_no_solution_cause_dict_end:
                    compare_no_solution_cause_dict_end[(key[0],key[1])] = 1
                else:
                    compare_no_solution_cause_dict_end[(key[0],key[1])] += 1
            if result == '1' and str(key[0])[0] != '0':
                print('result exsit')
                if 'status' not in result_dict:
                    result_dict['status'] = []
                    result_dict['status'].append('noSolutionOptimal')
                else:
                    result_dict['status'].append('noSolutionOptimal')
                    
                if 'year' not in result_dict:
                    result_dict['year'] = []
                    result_dict['year'].append(year)
                else:
                    result_dict['year'].append(year)
                if 'month' not in result_dict:
                    result_dict['month'] = []
                    result_dict['month'].append(month)
                else:
                    result_dict['month'].append(month)
                if 'date' not in result_dict:
                    result_dict['date'] = []
                    result_dict['date'].append(day)
                else:
                    result_dict['date'].append(day)
                if 'shift' not in result_dict:
                    result_dict['shift'] = []
                    result_dict['shift'].append(shift)
                else:
                    result_dict['shift'].append(shift)
                if 'shift_code' not in result_dict:
                    result_dict['shift_code'] = []
                    result_dict['shift_code'].append(shift_code)
                else:
                    result_dict['shift_code'].append(shift_code)
                if 'line_id' not in result_dict:
                    result_dict['line_id'] = []
                    result_dict['line_id'].append(key[0])
                else:
                    result_dict['line_id'].append(key[0])
                if 'OJT_code'not in result_dict:
                    result_dict['OJT_code'] = []
                    result_dict['OJT_code'].append(key[1])
                else:
                    result_dict['OJT_code'].append(key[1])
                if 'person_no' not in result_dict:
                    result_dict['person_no'] = []
                    result_dict['person_no'].append(key[2])
                else:
                    result_dict['person_no'].append(key[2])   
            elif result == '1' and str(key[0])[0] == '0':
                line_list = filter_line_group_to_one_person_dict[key[1]][key[0][0]+'.'+key[0][1:]]['line']
                for l in line_list:
                    if 'status' not in result_dict:
                        result_dict['status'] = []
                        result_dict['status'].append('noSolutionOptimal')
                    else:
                        result_dict['status'].append('noSolutionOptimal')
                    if 'year' not in result_dict:
                        result_dict['year'] = []
                        result_dict['year'].append(year)
                    else:
                        result_dict['year'].append(year)
                    if 'month' not in result_dict:
                        result_dict['month'] = []
                        result_dict['month'].append(month)
                    else:
                        result_dict['month'].append(month)
                    if 'date' not in result_dict:
                        result_dict['date'] = []
                        result_dict['date'].append(day)
                    else:
                        result_dict['date'].append(day)
                    if 'shift' not in result_dict:
                        result_dict['shift'] = []
                        result_dict['shift'].append(shift)
                    else:
                        result_dict['shift'].append(shift)
                    if 'shift_code' not in result_dict:
                        result_dict['shift_code'] = []
                        result_dict['shift_code'].append(shift_code)
                    else:
                        result_dict['shift_code'].append(shift_code)
                    if 'line_id' not in result_dict:
                        result_dict['line_id'] = []
                        result_dict['line_id'].append(l)
                    else:
                        result_dict['line_id'].append(l)
                    if 'OJT_code'not in result_dict:
                        result_dict['OJT_code'] = []
                        result_dict['OJT_code'].append(key[1])
                    else:
                        result_dict['OJT_code'].append(key[1])
                    if 'person_no' not in result_dict:
                        result_dict['person_no'] = []
                        result_dict['person_no'].append(key[2])
                    else:
                        result_dict['person_no'].append(key[2])
                
        return 'noSolutionOptimal', 'second round optimal solution found in %s - %s of group (%s) on %s/%s/%s'  % (shift,shift_code,working_line_name_list_string,year,month,day), result_dict,compare_no_solution_cause_dict_end


    else:
        print('*** No solution found in round two ***')
        return 'nosolution','second round still no solution found in %s - %s of group (%s) on %s/%s/%s' % (shift,shift_code,working_line_name_list_string,year,month,day),[],{}


# In[22]:


def find_off_person_used_in_result_df(merged_part_result_df):
    off_person_used_df = merged_part_result_df[merged_part_result_df['SAPShiftCode']=='OFF']
    off_person_used_list = list(set(off_person_used_df['PersonNo']))
    return off_person_used_list


# In[23]:


def main():
    #线体相关基础数据
    line_id_dict = read_excel_line_id(r'.\Data\autoshift.xlsx')
    working_line_dict = read_excel_working_line()
    line_skill_prefer_regular_dict = read_excel_line_need_skill_hc(r'.\Data\autoshift.xlsx')
    #人员相关基础数据
    ll_info_df, op_info_df = read_excel_person_info(r'.\Data\HC.xlsx')
    ot_df = read_excel_ishopfloor_person_ot(r'.\Data\ishopfloor_front_OT.xlsx')
    tqs_df = read_excel_tqs_skill(r'.\Data\1127_TQS.xlsx')
    ishopfloor_shift_df = read_excel_ishopfloor_person_shift(r'.\Data\ishopfloor_end_shift.xlsx')
    
    off_person_used_list = []
    result_description_list =[]
    result_df_list = []
    no_solution_cause_df_list = []
    count =0
    for key,value in working_line_dict.items():
        for line_group_list in value:
            count+=1
            year = key[0]
            month = key[1]
            day = key[2]
            shift = key[3]
            shift_code = key[4]
            working_line_id_list = line_group_list
            working_line_name_list = []
            for line_id in working_line_id_list:
                working_line_name_list.append(get_keys_from_value(line_id_dict, line_id))
            working_line_name_list_string = ', '.join(working_line_name_list)
            
            #working line需要的人头 （包含group用0开头的表示）    
            filter_line_group_to_one_person_dict, line_skill_perfer_to_hc_dict  = working_line_prefer_regular_hc(line_skill_prefer_regular_dict,working_line_id_list)
            print(filter_line_group_to_one_person_dict)
            print('==========================================%s %s %s %s %s %s=================================' %(year,month,day,shift,shift_code,working_line_name_list_string))
            print('*********line_skill_perfer_to_hc_dict ',line_skill_perfer_to_hc_dict)
#             new_dict = {}
#             for key2,value2 in line_skill_perfer_to_hc_dict.items():
#                 if str(key2[0])[0] == '0':
#                     line_list = filter_line_group_to_one_person_dict[key2[1]][key2[0][0]+'.'+key2[0][1:]]['line']
#                     for l in line_list:
#                         new_dict[(l,key2[1],key2[2])] = key2[0][0]+'.'+key2[0][1:]
#                 else:
#                     new_dict[key2] = value2
                    
#             line_base_data_df = pd.DataFrame(list(new_dict.items()),columns=['Key', 'Value'])
#             line_base_data_df.to_excel(r'.\line_base_data_%s.xlsx' %(str(count)),index=False)
           
            
            #人员
            person_for_use_df = person_for_use(op_info_df,ot_df,tqs_df,ishopfloor_shift_df,year,month,day,shift,shift_code,line_id_dict,working_line_id_list,off_person_used_list)
            print('***********person_for_use_df ',person_for_use_df)
#             person_base_data_df = pd.DataFrame(person_for_use_df)
#             person_base_data_df.to_excel(r'.\person_base_data_%s.xlsx' %(str(count)),index=False)
#             print(person_for_use_df[person_for_use_df['OJTCode'] == 'OJT3262'])
            status,description,part_result_dict = auto_shift(year,month,day,shift,shift_code,line_skill_perfer_to_hc_dict,filter_line_group_to_one_person_dict,person_for_use_df,working_line_name_list_string)  
            
            if status == 'optimal':
                result_description_list.append(description)
                part_result_df = pd.DataFrame(part_result_dict)
                line_list = part_result_df['line_id'].apply(lambda x: get_keys_from_value(line_id_dict, x))
                part_result_df.insert(6, 'line', line_list)
                part_result_df.rename(columns={'status':'Status','year':'Year','month':'Month','date':'Day','shift':'Shift','shift_code':'ShiftCode','line_id':'LineId','line':'Line','OJT_code':'OJTCode','person_no':'PersonNo'},inplace=True)
                part_result_df['PersonNo'] = part_result_df['PersonNo'].astype(int).astype(str)    
                person_info_for_result_df = person_for_use_df[['PersonNo','Name','OJTCode','SkillName','Sshift','sap_shift_code','年度AveNew','当月OT','性别']]
                person_info_for_result_df.rename(columns={'sap_shift_code':'SAPShiftCode','年度AveNew':'ytdAveNewOT','当月OT':'MonthOT','性别':'Gender'},inplace=True)
                
                merged_part_result_df = pd.merge(part_result_df, person_info_for_result_df, on=['PersonNo', 'OJTCode'], how='left')
                merged_part_result_df.drop_duplicates(inplace=True)
                result_df_list.append(merged_part_result_df)
    
                #按排班顺序累积用过的offperson保证下一次不会再用到
                off_person_used_list += find_off_person_used_in_result_df(merged_part_result_df)
            else:
                result_description_list.append(description)
                status,description,part_result_dict,compare_no_solution_cause_dict_end = auto_shift_no_solution(year,month,day,shift,shift_code,line_skill_perfer_to_hc_dict,filter_line_group_to_one_person_dict,person_for_use_df,working_line_name_list_string)
                if status == 'noSolutionOptimal':
                    result_description_list.append(description)
                    # 比对本轮所需线体技能人数 和 实际配上人员的线体技能人数 来求出哪些位置是空缺的
                    print('=======================compare_no_solution_cause_dict_end ',compare_no_solution_cause_dict_end)
                    #本轮所需线体技能人数的字典去掉key里面的第三位gender元素，保留（线体id，技能OJTCODE)
                    compare_no_solution_cause_dict_front = {}
                    for line_skill_perfer_to_hc_dict_key, line_skill_perfer_to_hc_dict_value in line_skill_perfer_to_hc_dict.items():
                        compare_no_solution_cause_dict_front[(line_skill_perfer_to_hc_dict_key[0],line_skill_perfer_to_hc_dict_key[1])] = line_skill_perfer_to_hc_dict_value
                    #实际安排人了的字典
                    print('=======================compare_no_solution_cause_dict_front ',compare_no_solution_cause_dict_front)
                    # 找到两个字典共有的键集合
                    common_keys = set(compare_no_solution_cause_dict_front.keys()) & set(compare_no_solution_cause_dict_end.keys())
                    # 对共有的键进行值的相减操作
                    part_no_solution_cause_dict = {common_key: compare_no_solution_cause_dict_front[common_key] - compare_no_solution_cause_dict_end[common_key] for common_key in common_keys}
                    # 将 compare_no_solution_cause_dict_front 中剩余的键加入到结果字典中
                    for front_key in set(compare_no_solution_cause_dict_front.keys()) - set(compare_no_solution_cause_dict_end.keys()):
                        part_no_solution_cause_dict[front_key] = compare_no_solution_cause_dict_front[front_key]
                    
                    # Remove key-value pairs where the value is 0
                    filter_part_no_solution_cause_dict = {part_no_solution_cause_dict_key: part_no_solution_cause_dict_value for part_no_solution_cause_dict_key, part_no_solution_cause_dict_value in part_no_solution_cause_dict.items() if part_no_solution_cause_dict_value != 0}
                    
                    
                    part_no_solution_cause_df = pd.DataFrame(list(filter_part_no_solution_cause_dict.items()),columns=['Key', 'Value'])
                    part_no_solution_cause_df['Year'] = year
                    part_no_solution_cause_df['Month'] = month
                    part_no_solution_cause_df['Date'] = day
                    part_no_solution_cause_df['Shift'] = shift
                    part_no_solution_cause_df['ShiftCode'] = shift_code
#                     part_no_solution_cause_df['Line'] = part_no_solution_cause_df['Key'].apply(lambda x: x if x[0][0] == '0' else get_keys_from_value(line_id_dict, int(x)))
                    
                    no_solution_cause_df_list.append(part_no_solution_cause_df)
                    
                    #把有解的放到结果里
                    if part_result_dict:
                        part_result_df = pd.DataFrame(part_result_dict)
                        line_list = part_result_df['line_id'].apply(lambda x: get_keys_from_value(line_id_dict, x))
                        part_result_df.insert(6, 'line', line_list)
                        part_result_df.rename(columns={'status':'Status','year':'Year','month':'Month','date':'Day','shift':'Shift','shift_code':'ShiftCode','line_id':'LineId','line':'Line','OJT_code':'OJTCode','person_no':'PersonNo'},inplace=True)
                        part_result_df['PersonNo'] = part_result_df['PersonNo'].astype(int).astype(str)    
                        person_info_for_result_df = person_for_use_df[['PersonNo','Name','OJTCode','SkillName','Sshift','sap_shift_code','年度AveNew','当月OT','性别']]
                        person_info_for_result_df.rename(columns={'sap_shift_code':'SAPShiftCode','年度AveNew':'ytdAveNewOT','当月OT':'MonthOT','性别':'Gender'},inplace=True)

                        merged_part_result_df = pd.merge(part_result_df, person_info_for_result_df, on=['PersonNo', 'OJTCode'], how='left')
                        merged_part_result_df.drop_duplicates(inplace=True)
                        result_df_list.append(merged_part_result_df)
                        
                        #按排班顺序累积用过的offperson保证下一次不会再用到
                        off_person_used_list += find_off_person_used_in_result_df(merged_part_result_df)
                                      

    result_df = pd.concat(result_df_list)
    result_df.to_excel(r'.\auto_shift_result.xlsx',index=False)
    df_to_sql(result_df)
    if no_solution_cause_df_list:
        no_solution_cause_df = pd.concat(no_solution_cause_df_list)
        no_solution_cause_df.to_excel(r'.\auto_shift_no_solution.xlsx',index=False)
    description_df = pd.DataFrame({'Description':result_description_list})
    description_df.to_excel(r'.\auto_shift_description.xlsx',index=False)
